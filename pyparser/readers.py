# -*- coding: utf-8 -*-
"""Чтение исходных файлов инвойсов в единый вид.

Каждый ридер возвращает объект Doc:
  text  — весь текст файла одной строкой (для поиска по меткам),
  rows  — список строк-ячеек: List[List[str]] (для xls/xlsx — ячейки листа,
          для pdf — построчный текст),
  meta  — служебная информация (кол-во страниц/листов, есть ли текст).
"""
from __future__ import annotations
import os
import warnings
import logging
from dataclasses import dataclass, field
from typing import List

warnings.filterwarnings("ignore")
logging.disable(logging.CRITICAL)


@dataclass
class Doc:
    path: str
    kind: str                       # 'pdf' | 'xls' | 'xlsx'
    rows: List[List[str]] = field(default_factory=list)
    text: str = ""
    meta: dict = field(default_factory=dict)

    @property
    def basename(self) -> str:
        return os.path.basename(self.path)


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def read_pdf(path: str) -> Doc:
    import pdfplumber
    rows: List[List[str]] = []
    parts: List[str] = []
    npages = 0
    with pdfplumber.open(path) as pdf:
        npages = len(pdf.pages)
        for p in pdf.pages:
            t = p.extract_text() or ""
            parts.append(t)
            for line in t.splitlines():
                rows.append([line])
    text = "\n".join(parts)
    meta = {"pages": npages, "has_text": len(text.strip()) >= 40, "ocr": False}
    if not meta["has_text"]:                         # скан — пробуем OCR
        try:
            from . import ocr
            if ocr.available():
                otext = ocr.ocr_pdf(path)
                if len(otext.strip()) >= 40:
                    text = otext
                    rows = [[l] for l in otext.splitlines()]
                    meta["has_text"] = True
                    meta["ocr"] = True
        except Exception as e:
            meta["ocr_error"] = repr(e)
    return Doc(path, "pdf", rows, text, meta)


def _plausible(d) -> bool:
    import datetime
    return isinstance(d, (datetime.date, datetime.datetime)) and 2023 <= d.year <= 2028


def read_xlsx(path: str) -> Doc:
    import openpyxl
    import datetime
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: List[List[str]] = []
    parts: List[str] = []
    cell_dates = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            for v in r:
                if isinstance(v, (datetime.date, datetime.datetime)) and _plausible(v):
                    cell_dates.append(v.date() if isinstance(v, datetime.datetime) else v)
            cells = [_cell_str(v) for v in r]
            if any(c.strip() for c in cells):
                rows.append(cells)
                parts.append(" ".join(c for c in cells if c.strip()))
    wb.close()
    text = "\n".join(parts)
    return Doc(path, "xlsx", rows, text,
               {"has_text": len(text.strip()) >= 20, "cell_dates": cell_dates})


class _DevNull:
    def write(self, *a):
        pass


def read_xls(path: str) -> Doc:
    import xlrd
    book = xlrd.open_workbook(path, logfile=_DevNull())
    rows: List[List[str]] = []
    parts: List[str] = []
    cell_dates = []
    for sh in book.sheets():
        for r in range(sh.nrows):
            cells = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == 3:  # XL_CELL_DATE
                    try:
                        d = xlrd.xldate_as_datetime(cell.value, book.datemode).date()
                        if _plausible(d):
                            cell_dates.append(d)
                    except Exception:
                        pass
                cells.append(_cell_str(cell.value))
            if any(c.strip() for c in cells):
                rows.append(cells)
                parts.append(" ".join(c for c in cells if c.strip()))
    text = "\n".join(parts)
    return Doc(path, "xls", rows, text,
               {"sheets": book.nsheets, "has_text": len(text.strip()) >= 20,
                "cell_dates": cell_dates})


def read_any(path: str) -> Doc:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return read_pdf(path)
    if ext == ".xlsx":
        return read_xlsx(path)
    if ext == ".xls":
        return read_xls(path)
    raise ValueError("unsupported file type: " + ext)
