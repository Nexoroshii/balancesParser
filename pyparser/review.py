# -*- coding: utf-8 -*-
"""Редактируемая очередь ручной проверки.

export_queue() — пишет xlsx с нераспознанными файлами: подсказки (поставщик,
дата, суммы/кандидаты) уже заполнены, где известно; пользователь дописывает
недостающее и ставит «да» в колонке «Записывать».

read_queue() — читает заполненную очередь и возвращает строки к записи.

Затем `python -m pyparser.review --balance <файл> --queue <очередь.xlsx>` дописывает
эти строки в баланс (той же логикой BalanceWriter, с защитой от дублей).
"""
from __future__ import annotations
import argparse
import datetime as dt
import os
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .balance import BalanceWriter, norm
from .aliases import NEW_COLUMNS

HEADERS = ["Файл", "Поставщик (колонка баланса)", "Дата (ГГГГ-ММ-ДД)",
           "Суммы (через +)", "Записывать (да/нет)", "Подсказка / кандидаты"]

_HINT_FILL = PatternFill("solid", fgColor="FFF2CC")     # жёлтый — заполнить
_HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")


def export_queue(items: List[dict], path: str):
    """items: [{file, supplier, date(date|None), amounts[list], hint}]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "очередь"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for it in items:
        supplier = it.get("supplier") or ""
        date = it.get("date")
        amounts = it.get("amounts") or []
        formula = "+".join(_fmt(a) for a in amounts) if amounts else ""
        row = [
            it["file"],
            supplier,
            date.isoformat() if isinstance(date, dt.date) else "",
            formula,
            "",                       # Записывать — пользователь ставит «да»
            it.get("hint", ""),
        ]
        ws.append(row)
        r = ws.max_row
        # подсвечиваем ячейки, которые надо заполнить (пустой поставщик/сумма)
        if not supplier:
            ws.cell(r, 2).fill = _HINT_FILL
        if not formula:
            ws.cell(r, 4).fill = _HINT_FILL
        ws.cell(r, 5).fill = _HINT_FILL
    widths = [46, 30, 18, 26, 18, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def _fmt(x) -> str:
    x = float(x)
    return str(int(x)) if x.is_integer() else str(round(x, 2))


def _parse_amounts(s) -> List[float]:
    out = []
    for part in str(s).replace(" ", "").lstrip("=").split("+"):
        if not part:
            continue
        try:
            out.append(float(part.replace(",", ".")))
        except ValueError:
            return []
    return out


def _parse_date(s) -> Optional[dt.date]:
    if isinstance(s, dt.datetime):
        return s.date()
    if isinstance(s, dt.date):
        return s
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def read_queue(path: str) -> List[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["очередь"] if "очередь" in wb.sheetnames else wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        write_flag = str(ws.cell(r, 5).value or "").strip().lower()
        if write_flag not in ("да", "yes", "y", "1", "+", "true"):
            continue
        supplier = str(ws.cell(r, 2).value or "").strip()
        amounts = _parse_amounts(ws.cell(r, 4).value)
        date = _parse_date(ws.cell(r, 3).value)
        if not supplier or not amounts or not date:
            rows.append({"file": ws.cell(r, 1).value, "error":
                         "не заполнено: поставщик/сумма/дата"})
            continue
        rows.append({"file": ws.cell(r, 1).value, "supplier": supplier,
                     "date": date, "amounts": amounts})
    return rows


def apply_queue(balance: str, queue: str, out: str):
    from collections import defaultdict
    entries = read_queue(queue)
    w = BalanceWriter(balance, out)
    written, skipped, failed = [], [], []

    # агрегируем по поставщику (одна строка на поставщика; дата — самая ранняя)
    agg = defaultdict(lambda: {"amounts": [], "dates": []})
    for e in entries:
        if e.get("error"):
            failed.append((e["file"], e["error"]))
            continue
        agg[e["supplier"]]["amounts"].extend(e["amounts"])
        agg[e["supplier"]]["dates"].append(e["date"])

    for sup, rec in agg.items():
        if w.find_block(sup) is None and sup.upper() in {s.upper() for s in NEW_COLUMNS}:
            w.create_block(sup)
        status, msg = w.write_supplier(sup, min(rec["dates"]), rec["amounts"])
        (written if status == "written" else skipped if status == "skipped"
         else failed).append((sup, msg))
    w.save()
    return written, skipped, failed


def _auto(pattern, exclude=None):
    import glob
    cands = [p for p in glob.glob(os.path.join("excel", pattern))
             if not exclude or exclude not in os.path.basename(p)]
    return max(cands, key=os.path.getmtime) if cands else None


def main(argv=None):
    ap = argparse.ArgumentParser(description="Дописать заполненную очередь в баланс")
    ap.add_argument("--balance", default=None, help="файл баланса (по умолч. свежий *_auto.xlsx)")
    ap.add_argument("--queue", default=None, help="очередь (по умолч. свежая *_очередь.xlsx)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args(argv)
    import sys
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    print("=" * 60, flush=True)
    print("ДОПИСАТЬ ОЧЕРЕДЬ В БАЛАНС — старт", flush=True)
    print("=" * 60, flush=True)
    args.balance = args.balance or _auto("*_auto.xlsx")
    args.queue = args.queue or _auto("*_очередь.xlsx")
    if not args.balance or not args.queue:
        raise SystemExit("Не найден *_auto.xlsx или *_очередь.xlsx в excel/. Укажите --balance и --queue.")
    print("Баланс:  " + args.balance, flush=True)
    print("Очередь: " + args.queue, flush=True)
    print("Читаю очередь и дописываю в баланс...", flush=True)
    out = args.out or (os.path.splitext(args.balance)[0] + "_ready.xlsx")
    written, skipped, failed = apply_queue(args.balance, args.queue, out)
    print("Дописано из очереди: %d, пропущено(дубли): %d, ошибок: %d"
          % (len(written), len(skipped), len(failed)))
    for s, m in written:
        print("  +", s, "|", m)
    for s, m in failed:
        print("  ! ", s, "|", m)
    print("Готово:", out)


if __name__ == "__main__":
    main()
