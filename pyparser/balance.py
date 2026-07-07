# -*- coding: utf-8 -*-
"""Запись сумм инвойсов в файл «балансы актуальные» (Лист1).

Структура листа: горизонтальные блоки поставщиков, по 4 столбца
(дата | отгрузка | оплата | итого) + столбец-разделитель. Заголовок блока
(имя поставщика) — в строке 1 над столбцом «дата». Данные с 3-й строки.

Правила записи (см. память проекта):
- одна строка на поставщика за поставку;
- «отгрузка» = формула «=инв1+инв2+...» (суммы отдельных инвойсов);
- «итого» — накопительная формула, пересчитается сама;
- пишем в КОПИЮ, оригинал не трогаем.
"""
from __future__ import annotations
import re
import shutil
import datetime as dt
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

SHEET = "Лист1"
FIRST_DATA_ROW = 3


DEFAULT_DATE_FMT = "DD.MM.YYYY"


def norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip().casefold()


def _cell_date(v):
    """Дата из ячейки: datetime/date как есть, число-серийник Excel -> дата."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)) and 40000 <= v <= 50000:
        return dt.date(1899, 12, 30) + dt.timedelta(days=int(v))
    return None


@dataclass
class Block:
    name: str          # заголовок как в файле
    date_col: int      # столбец «дата»
    ship_col: int      # «отгрузка»
    pay_col: int       # «оплата»
    tot_col: int       # «итого»


class BalanceWriter:
    def __init__(self, src_path: str, out_path: str):
        self.src_path = src_path
        self.out_path = out_path
        shutil.copyfile(src_path, out_path)          # работаем на копии
        self.wb = openpyxl.load_workbook(out_path)   # с формулами
        self.ws = self.wb[SHEET]
        self.blocks: Dict[str, List[Block]] = {}
        self._index_blocks()

    def _index_blocks(self):
        ws = self.ws
        for c in range(1, ws.max_column + 1):
            name = ws.cell(1, c).value
            if name is None or not str(name).strip():
                continue
            # это заголовок блока, если под ним в строке 2 стоит «дата»
            sub = norm(ws.cell(2, c).value)
            if sub != norm("дата"):
                continue
            blk = Block(str(name), c, c + 1, c + 2, c + 3)
            self.blocks.setdefault(norm(name), []).append(blk)

    def _last_date(self, blk: Block) -> Optional[dt.date]:
        ws = self.ws
        latest = None
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            v = _cell_date(ws.cell(r, blk.date_col).value)
            if v and (latest is None or v > latest):
                latest = v
        return latest

    def find_block(self, canonical: str) -> Optional[Block]:
        lst = self.blocks.get(norm(canonical))
        if not lst:
            return None
        if len(lst) == 1:
            return lst[0]
        # несколько блоков с одним заголовком (напр. PACIFIC/FreshFlowLogistic):
        # берём тот, где последняя запись СВЕЖЕЕ (по правилу пользователя)
        return max(lst, key=lambda b: (self._last_date(b) or dt.date.min))

    # сколько ПОДРЯД пустых строк (по дата/отгрузка/оплата) считаем концом
    # живого массива данных: всё, что идёт ПОСЛЕ такого разрыва, — это посторонние
    # заметки/мусор в тех же столбцах («от Тесса», «коррект сумм», «???», числа без
    # строки и т.п.), дописывать после них нельзя.
    # Порог 8: в реальных данных встречаются внутренние разрывы до 3 пустых строк
    # (напр. EDANA FLORAL: строки 5-7 пусты, данные продолжаются), а мусор внизу
    # отделён ≥9 пустыми строками. 8 — безопасно между ними; ранняя остановка хуже
    # (пишет ВНУТРЬ блока, до существующих данных), поэтому берём ближе к верху.
    _GAP_STOP = 8

    def _next_empty_row(self, blk: Block) -> int:
        """Дописываем ПОСЛЕ последней строки живого массива данных.

        Живые данные идут сплошным массивом (строки только с «оплатой» — тоже
        данные, поэтому «первую пустую» брать нельзя). Первый разрыв в _GAP_STOP
        подряд пустых строк считаем концом блока — мусор ниже игнорируем."""
        ws = self.ws
        last = FIRST_DATA_ROW - 1
        gap = 0
        seen = False
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            filled = any(
                (v is not None and str(v).strip() != "")
                for v in (ws.cell(r, blk.date_col).value,
                          ws.cell(r, blk.ship_col).value,
                          ws.cell(r, blk.pay_col).value)
            )
            if filled:
                last = r
                gap = 0
                seen = True
            elif seen:
                gap += 1
                if gap >= self._GAP_STOP:
                    break
        return last + 1

    def _date_format_from_above(self, blk: Block, row: int) -> str:
        """Формат даты из ближайшей заполненной ячейки выше; если он не датовый
        (напр. General/число — тогда дата покажется как 46200) — берём дефолт."""
        for r in range(row - 1, FIRST_DATA_ROW - 1, -1):
            c = self.ws.cell(r, blk.date_col)
            if c.value is not None:
                fmt = c.number_format or ""
                if re.search(r"[dmyDMY]", fmt) and "General" not in fmt:
                    return fmt
                break
        return DEFAULT_DATE_FMT

    def _existing_row(self, blk: Block, date: dt.date, formula: str) -> Optional[int]:
        """Ищет уже записанную строку той же поставки (дата + формула отгрузки)."""
        ws = self.ws
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            s = ws.cell(r, blk.ship_col).value
            if s is None or str(s).strip() != formula:
                continue
            if _cell_date(ws.cell(r, blk.date_col).value) == date:
                return r
        return None

    def write_supplier(self, canonical: str, date: dt.date,
                       amounts: List[float]) -> Tuple[str, str]:
        """Записывает одну строку. Возвращает (статус, сообщение).
        статус: 'written' | 'skipped' | 'error'."""
        blk = self.find_block(canonical)
        if blk is None:
            return "error", "блок поставщика не найден в балансе"

        formula = "=" + "+".join(_fmt_num(a) for a in amounts)
        # защита от повторной записи той же поставки
        dup = self._existing_row(blk, date, formula)
        if dup is not None:
            return "skipped", f"уже записано (строка {dup})"

        row = self._next_empty_row(blk)
        ws = self.ws

        # дата
        dcell = ws.cell(row, blk.date_col)
        dcell.value = dt.datetime(date.year, date.month, date.day)
        dcell.number_format = self._date_format_from_above(blk, row)

        # отгрузка = формула =a+b+c
        ws.cell(row, blk.ship_col).value = formula

        # итого = накопление (если формулы ещё нет в этой строке — добавим)
        tot = ws.cell(row, blk.tot_col)
        if tot.value is None or str(tot.value).strip() == "":
            tl = get_column_letter(blk.tot_col)
            sl = get_column_letter(blk.ship_col)
            pl = get_column_letter(blk.pay_col)
            if row - 1 >= FIRST_DATA_ROW:
                tot.value = f"={tl}{row-1}-{sl}{row}+{pl}{row}"
            else:                                    # первая строка блока — предыдущего итого нет
                tot.value = f"=-{sl}{row}+{pl}{row}"

        col = get_column_letter(blk.date_col)
        return "written", f"строка {row} (кол. {col}): {formula}"

    def create_block(self, name: str) -> Block:
        """Создаёт новый блок поставщика в конце листа (в стиле остальных)."""
        ws = self.ws
        # ищем последний занятый столбец по строке 2 (подзаголовки)
        last = 1
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value or ws.cell(2, c).value:
                last = c
        start = last + 2                     # +1 разделитель
        ws.cell(1, start).value = name
        for i, sub in enumerate(("дата", "отгрузка", "оплата", "итого")):
            ws.cell(2, start + i).value = sub
        blk = Block(name, start, start + 1, start + 2, start + 3)
        self.blocks.setdefault(norm(name), []).append(blk)
        # начальное «итого» для первой строки данных не задаём — проставит write_supplier
        return blk

    def save(self):
        self.wb.save(self.out_path)


def _fmt_num(x: float) -> str:
    if float(x).is_integer():
        return str(int(x))
    return str(round(x, 2))
